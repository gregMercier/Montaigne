
        
# -*- coding: utf-8 -*-
"""
Created on Wed Feb 28 16:37:45 2018

@author: GREGORYMercier
"""
import pandas as pd
from facepy import GraphAPI
import openpyxl
 

#######################identification#######################################

App_id = ""
App_secret = ""
access_token = App_id + "|" + App_secret

#########################retrieve since date##################################


wb = openpyxl.load_workbook('since_date.xlsx')        
sheet = wb.get_sheet_by_name('since_date')
local='A'+str(sheet.max_row)
since_inp= sheet[local].value  
since_inp=str(since_inp)
 

##############################################################################


until_inp =  strftime("%Y-%m-%d")
since = "since=" + since_inp
until = "until=" + str(until_inp)
node = page_id + "/"+"posts?"+ since+"&"+until+"&"
parameters = "fields=message,created_time&access_token=%s" % ( access_token)

url = node + parameters

graph = GraphAPI(access_token)


i=0
posts=[]
for page_id in list_url:
        node = page_id + "/"+"posts?"+ since+"&"+until+"&"
        url = node + parameters
        globals()["datas" + str(i)]=graph.get(url, page=True, retry=5)
        for data in globals()["datas" + str(i)]:
              posts.append(data)
              i = i+1


#############################################extract key- values 'data'########
#################################from the list of dictionnary posts############

posts_data=[d['data'] for d in posts]


#########################transform list of list of dictionnary################
##### into a list of dictionnary post named message ##########################
post_message=[]
i=0
while i <len(posts)+1: 
    post_message.extend(posts_data[i])
    i =i+1

#########################################################################

######################Drop all the posts without message ######################
################### keep them in post_message_only#############################
post_message_only=[]

i=0
while i <len(post_message):
    if (len(post_message[i]) == 3):
        post_message_only.append(post_message[i])
        i =i+1
    else:
        i =i+1

##############################################################################


###########transform a list of dictionary into a list of values ##############
##############################################################################

post_message5 = [[x['id'],x['created_time'],x['message']] for x in post_message_only]

#############################################################################


##############################creation du fichier initial ####################



my_df = pd.DataFrame(post_message5)

my_df.to_csv('my_csv_file.csv',index=False,header=False)


####################ajout des noms de colonnes ################################


df= pd.read_csv('my_csv_file.csv', sep=',', encoding='latin-1')

df.columns = ['user_id','date_created','message']

#############separate user_id into user_id and id_message #####################

i = df.columns.get_loc('user_id')
df2 = df['user_id'].str.split("_", expand=True)

df3= pd.concat([df.iloc[:, :i], df2, df.iloc[:, i+1:]], axis=1)

df3.columns = ['user_id','id_message','date_created','message']

df3.to_csv('extractfrom'+strftime("%Y-%m-%d")+'.csv',index=False)

###############################################################################


list_url = [
'163293133759468',
'192953374168500',
'262969587120491',
'154409134651718',
'121470551304281',
'12619074644',
'350292144987173',
'192116037510649',
'59627025578',
'527856500606836',
'391396000937067',
'642045892502440',

];
        
################insert MAJ date in the excel file############################
#############################################################################

wb = openpyxl.load_workbook('since_date.xlsx')
sheet = wb.get_sheet_by_name('since_date')
new_date='A'+str(sheet.max_row+1)
sheet[new_date] = strftime("%Y-%m-%d")
wb.save('since_date.xlsx')


##############################################################################
##############################################################################